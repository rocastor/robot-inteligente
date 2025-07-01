"""
🤖 Módulo de Generadores de Archivos
Maneja la creación de JSON, PDF, Excel y otros formatos, con soporte para Google Drive.
"""

import json
import os
import shutil
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# Configuración de almacenamiento - Google Drive como principal
DRIVE_AVAILABLE = True  # Google Drive como almacenamiento principal
S3_AVAILABLE = False    # S3 deshabilitado - usamos solo Google Drivele Drive

def guardar_json(data, nombre_base):
    """Guarda datos en formato JSON con soporte para Google Drive"""
    try:
        nombre_archivo = f"{os.path.basename(nombre_base)}.json"
        proceso_nombre = data.get('metadatos_proceso', {}).get('id_unico_proceso', 'unknown')
        carpeta_original = data.get('metadatos_proceso', {}).get('carpeta_original_detectada', '')

        temp_file = f"/tmp/{nombre_archivo}"
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"📄 JSON creado temporalmente: {temp_file}")

        # Verificar disponibilidad de Google Drive
        if not DRIVE_AVAILABLE:
            print("⚠️ Google Drive no está disponible - configurar credenciales")
            print("💡 Para usar Google Drive, configure GOOGLE_DRIVE_CREDENTIALS en Secrets")
            raise Exception("Google Drive no configurado. Configure GOOGLE_DRIVE_CREDENTIALS para continuar.")
        else:
            print("☁️ Intentando usar Google Drive como almacenamiento principal")

        try:
            from modules.google_drive_client import get_drive_client
            drive_client = get_drive_client()
            if not drive_client or not drive_client.service:
                raise Exception("No se pudo conectar a Google Drive - verificar credenciales")

            # Obtener metadatos del proceso
            metadatos = data.get('metadatos_proceso', {})
            carpeta_original = metadatos.get('carpeta_original_detectada', 'Sin_Carpeta')
            nombre_proceso = metadatos.get('id_unico_proceso', 'proceso_unknown')

            # Crear estructura de carpetas en Drive
            main_folder_id = drive_client.create_or_get_folder('Robot_AI_Procesos')
            original_folder_id = drive_client.create_or_get_folder(carpeta_original, main_folder_id)
            process_folder_id = drive_client.create_or_get_folder(nombre_proceso, original_folder_id)

            # Subir archivo
            metadata = {
                'tipo': 'analisis_json',
                'proceso': nombre_proceso,
                'carpeta_original': carpeta_original,
                'timestamp': datetime.now().isoformat()
            }

            drive_result = drive_client.upload_file(
                temp_file, 
                nombre_archivo, 
                process_folder_id,
                metadata
            )

            if drive_result:
                print(f"☁️ JSON subido a Google Drive: {drive_result['web_view_link']}")

                # Agregar información de Drive al JSON
                data['google_drive_info'] = {
                    'file_id': drive_result['id'],
                    'web_view_link': drive_result['web_view_link'],
                    'folder_id': process_folder_id,
                    'upload_time': drive_result['upload_time']
                }

                # Actualizar archivo local con info de Drive
                with open(temp_file, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)

                print(f"✅ JSON procesado exitosamente con Google Drive")
            else:
                raise Exception("Error subiendo archivo a Google Drive")

        except Exception as e:
            print(f"❌ Error crítico con Google Drive: {str(e)}")
            raise Exception(f"Error guardando en Google Drive: {str(e)}")
        return temp_file
    except Exception as e:
        print(f"❌ Error guardando JSON: {str(e)}")
        return None

def guardar_pdf(data, nombre_base):
    """Genera un reporte PDF con los resultados del análisis"""
    try:
        nombre_archivo = f"{nombre_base}_reporte.pdf"

        styles = getSampleStyleSheet()
        titulo_style = ParagraphStyle(
            'TituloPersonalizado',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1
        )

        subtitulo_style = ParagraphStyle(
            'SubtituloPersonalizado',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=15,
            textColor='blue'
        )

        normal_style = ParagraphStyle(
            'NormalPersonalizado',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=10
        )

        story = []
        story.append(Paragraph("🤖 REPORTE DE ANÁLISIS AUTOMÁTICO", titulo_style))
        story.append(Spacer(1, 20))

        story.append(Paragraph("📊 RESUMEN EJECUTIVO", subtitulo_style))
        resumen = data.get('resumen', {})
        costos = data.get('costos_openai', {})

        story.append(Paragraph(f"• Archivos procesados: {resumen.get('archivos_procesados_exitosamente', 0)}/{resumen.get('archivos_recibidos', 0)}", normal_style))
        story.append(Paragraph(f"• Caracteres extraídos: {resumen.get('caracteres_totales_extraidos', 0):,}", normal_style))
        story.append(Paragraph(f"• Preguntas analizadas: {resumen.get('preguntas_analizadas', 0)}", normal_style))
        story.append(Paragraph(f"• Respuestas con información: {resumen.get('respuestas_con_informacion', 0)}", normal_style))

        if costos:
            story.append(Paragraph(f"• Costo total OpenAI: ${costos.get('costo_total_usd', 0):.4f} USD", normal_style))
            story.append(Paragraph(f"• Tokens utilizados: {costos.get('tokens_totales_usados', 0):,}", normal_style))
            story.append(Paragraph(f"• Modelo: {costos.get('modelo_utilizado', 'N/A')}", normal_style))

        story.append(Paragraph(f"• Fecha y hora: {data.get('timestamp', '')}", normal_style))
        story.append(Spacer(1, 20))

        story.append(Paragraph("📁 ARCHIVOS PROCESADOS", subtitulo_style))
        for archivo in data.get('archivos', []):
            estado = "✅ Exitoso" if archivo.get('procesado') else "❌ Error"
            story.append(Paragraph(f"• {archivo.get('nombre', 'Sin nombre')} - {estado}", normal_style))
            story.append(Paragraph(f"  Tipo: {archivo.get('tipo', 'Desconocido')} | Caracteres: {archivo.get('caracteres_extraidos', 0)}", normal_style))
        story.append(Spacer(1, 20))

        story.append(Paragraph("📋 DATOS EXTRAÍDOS", subtitulo_style))

        etiqueta_style = ParagraphStyle(
            'EtiquetaStyle',
            parent=styles['Normal'],
            fontSize=11,
            fontName='Helvetica-Bold',
            textColor='darkblue',
            spaceAfter=8
        )

        etiquetas_mapeo = {
            "¿Cuál es el nombre de la entidad?": "ENTIDAD:",
            "¿Cuál es el NIT de la entidad?": "NIT:",
            "¿Cuál es la dirección de la entidad?": "DIRECCIÓN:",
            "¿Cuál es la ciudad donde se encuentra la entidad?": "CIUDAD:",
            "¿Cuál es el objeto del contrato?": "OBJETO:",
            "¿Cuál es el valor del contrato?": "VALOR:",
            "¿Qué menciona de experiencia?": "EXPERIENCIA:",
            "¿Qué menciona sobre salud y pensión?": "SALUD Y PENSIÓN:",
            "¿Qué anexos o formatos se mencionan?": "ANEXOS/FORMATOS:",
            "¿Cuál es el cronograma?": "CRONOGRAMA:"
        }

        for resultado in data.get('analisis', []):
            pregunta = resultado.get('pregunta', '')
            respuesta = resultado.get('respuesta', '')

            etiqueta = etiquetas_mapeo.get(pregunta, f"{pregunta}:")

            if resultado.get('informacion_encontrada') and respuesta != "No se encontró información específica para esta pregunta":
                story.append(Paragraph(etiqueta, etiqueta_style))
                story.append(Paragraph(respuesta, normal_style))
                story.append(Spacer(1, 10))

        temp_pdf = f"/tmp/{nombre_archivo}"
        doc_temp = SimpleDocTemplate(temp_pdf, pagesize=letter, 
                              leftMargin=inch, rightMargin=inch,
                              topMargin=inch, bottomMargin=inch)
        doc_temp.build(story)
        print(f"📄 PDF creado temporalmente: {temp_pdf}")
        metadatos = data.get('metadatos_proceso', {})
        carpeta_original = metadatos.get('carpeta_original_detectada', 'Sin_Carpeta')
        nombre_proceso = metadatos.get('id_unico_proceso', 'proceso_unknown')

        # Verificar disponibilidad de Google Drive
        if not DRIVE_AVAILABLE:
            print("⚠️ Google Drive no está disponible - configurar credenciales")
            print("💡 Para usar Google Drive, configure GOOGLE_DRIVE_CREDENTIALS en Secrets")
            raise Exception("Google Drive no configurado. Configure GOOGLE_DRIVE_CREDENTIALS para continuar.")
        else:
            print("☁️ Intentando usar Google Drive como almacenamiento principal")

        # Subir a Google Drive (almacenamiento principal)
        try:
            from modules.google_drive_client import get_drive_client
            drive_client = get_drive_client()
            if not drive_client or not drive_client.service:
                raise Exception("No se pudo conectar a Google Drive - verificar credenciales")

            # Crear estructura de carpetas
            main_folder_id = drive_client.create_or_get_folder('Robot_AI_Procesos')
            original_folder_id = drive_client.create_or_get_folder(carpeta_original, main_folder_id)
            process_folder_id = drive_client.create_or_get_folder(nombre_proceso, original_folder_id)

            # Subir PDF
            metadata = {
                'tipo': 'reporte_pdf',
                'proceso': nombre_proceso,
                'carpeta_original': carpeta_original,
                'timestamp': datetime.now().isoformat()
            }

            drive_result = drive_client.upload_file(
                temp_pdf, 
                nombre_archivo, 
                process_folder_id,
                metadata
            )

            if drive_result:
                print(f"☁️ PDF subido a Google Drive: {drive_result['web_view_link']}")
            else:
                raise Exception("Error subiendo PDF a Google Drive")

        except Exception as e:
            print(f"❌ Error crítico subiendo PDF a Google Drive: {str(e)}")
            raise Exception(f"Error guardando PDF en Google Drive: {str(e)}")

        return temp_pdf

    except Exception as e:
        print(f"❌ Error generating PDF: {str(e)}")

def guardar_excel(data, nombre_base):
    """Genera un archivo Excel organizado con los resultados del análisis"""
    try:
        nombre_archivo = f"{nombre_base}_analisis.xlsx"
        wb = Workbook()

        # Estilos
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # HOJA 1: RESUMEN EJECUTIVO
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Ejecutivo"

        ws_resumen['A1'] = '🤖 ROCASTOR ROBOT - ANÁLISIS AUTOMÁTICO'
        ws_resumen['A1'].font = Font(name='Arial', size=16, bold=True)
        ws_resumen.merge_cells('A1:B1')

        resumen = data.get('resumen', {})
        costos = data.get('costos_openai', {})

        info_general = [
            ['📊 ESTADÍSTICAS GENERALES', ''],
            ['Fecha de Análisis', data.get('timestamp', '')],
            ['Archivos Procesados', f"{resumen.get('archivos_procesados_exitosamente', 0)}/{resumen.get('archivos_recibidos', 0)}"],
            ['Caracteres Extraídos', f"{resumen.get('caracteres_totales_extraidos', 0):,}"],
            ['Preguntas Analizadas', resumen.get('preguntas_analizadas', 0)],
            ['Respuestas Encontradas', resumen.get('respuestas_con_informacion', 0)],
            ['', ''],
            ['💰 COSTOS OPENAI', ''],
            ['Costo Total (USD)', f"${costos.get('costo_total_usd', 0):.4f}"],
            ['Tokens Utilizados', f"{costos.get('tokens_totales_usados', 0):,}"],
            ['Modelo Utilizado', costos.get('modelo_utilizado', 'N/A')],
            ['Costo por Pregunta', f"${costos.get('costo_promedio_por_pregunta', 0):.4f}"]
        ]

        for i, (label, value) in enumerate(info_general, 3):
            ws_resumen[f'A{i}'] = label
            ws_resumen[f'B{i}'] = value
            if label and not label.startswith('💰') and not label.startswith('📊'):
                ws_resumen[f'A{i}'].font = data_font
                ws_resumen[f'B{i}'].font = data_font
            elif label.startswith('💰') or label.startswith('📊'):
                ws_resumen[f'A{i}'].font = Font(name='Arial', size=11, bold=True)
                ws_resumen[f'A{i}'].fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

        # HOJA 2: DATOS EXTRAÍDOS
        ws_datos = wb.create_sheet("Datos Extraídos")

        columnas_mapeo = {
            "¿Cuál es el nombre oficial de la entidad contratante": "ENTIDAD",
            "¿Cuál es el número de NIT": "NIT", 
            "¿Cuál es la dirección física completa": "DIRECCIÓN",
            "¿En qué ciudad está ubicada": "CIUDAD",
            "¿Cuál es el objeto específico del contrato": "OBJETO",
            "¿Cuál es el valor total del contrato": "VALOR",
            "¿Qué requisitos de experiencia específica": "EXPERIENCIA",
            "¿Qué requisitos se mencionan sobre afiliación": "SALUD Y PENSIÓN",
            "¿Qué documentos anexos, formatos específicos": "ANEXOS/FORMATOS",
            "¿Cuál es el cronograma detallado": "CRONOGRAMA"
        }

        headers = ['PREGUNTA', 'RESPUESTA', 'INFO ENCONTRADA']
        for i, header in enumerate(headers, 1):
            cell = ws_datos.cell(row=1, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        row = 2
        for resultado in data.get('analisis', []):
            pregunta = resultado.get('pregunta', '')
            respuesta = resultado.get('respuesta', '')
            info_encontrada = "SÍ" if resultado.get('informacion_encontrada') else "NO"

            pregunta_corta = pregunta
            for key, value in columnas_mapeo.items():
                if key in pregunta:
                    pregunta_corta = value
                    break

            if resultado.get('informacion_encontrada') and respuesta != "No se encontró información específica para esta pregunta":
                ws_datos.cell(row=row, column=1, value=pregunta_corta).font = data_font
                ws_datos.cell(row=row, column=2, value=respuesta).font = data_font
                ws_datos.cell(row=row, column=3, value=info_encontrada).font = data_font

                for col in range(1, 4):
                    ws_datos.cell(row=row, column=col).border = border

                if row % 2 == 0:
                    for col in range(1, 4):
                        ws_datos.cell(row=row, column=col).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

                row += 1

        # Ajustar ancho de columnas
        for ws in [ws_resumen, ws_datos]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        temp_excel = f"/tmp/{nombre_archivo}"

        wb.save(temp_excel)
        print(f"📊 Excel creado temporalmente: {temp_excel}")

        # Verificar disponibilidad de Google Drive
        if not DRIVE_AVAILABLE:
            print("⚠️ Google Drive no está disponible - configurar credenciales")
            print("💡 Para usar Google Drive, configure GOOGLE_DRIVE_CREDENTIALS en Secrets")
            raise Exception("Google Drive no configurado. Configure GOOGLE_DRIVE_CREDENTIALS para continuar.")
        else:
            print("☁️ Intentando usar Google Drive como almacenamiento principal")

        # Subir a Google Drive (almacenamiento principal)
        try:
            from modules.google_drive_client import get_drive_client
            drive_client = get_drive_client()
            if not drive_client or not drive_client.service:
                raise Exception("No se pudo conectar a Google Drive - verificar credenciales")

            metadatos = data.get('metadatos_proceso', {})
            carpeta_original = metadatos.get('carpeta_original_detectada', 'Sin_Carpeta')
            nombre_proceso = metadatos.get('id_unico_proceso', 'proceso_unknown')

            # Crear estructura de carpetas
            main_folder_id = drive_client.create_or_get_folder('Robot_AI_Procesos')
            original_folder_id = drive_client.create_or_get_folder(carpeta_original, main_folder_id)
            process_folder_id = drive_client.create_or_get_folder(nombre_proceso, original_folder_id)

            # Subir Excel
            metadata = {
                'tipo': 'analisis_excel',
                'proceso': nombre_proceso,
                'carpeta_original': carpeta_original,
                'timestamp': datetime.now().isoformat()
            }

            drive_result = drive_client.upload_file(
                temp_excel, 
                nombre_archivo, 
                process_folder_id,
                metadata
            )

            if drive_result:
                print(f"☁️ Excel subido a Google Drive: {drive_result['web_view_link']}")
            else:
                raise Exception("Error subiendo Excel a Google Drive")

        except Exception as e:
            print(f"❌ Error crítico subiendo Excel a Google Drive: {str(e)}")
            raise Exception(f"Error guardando Excel en Google Drive: {str(e)}")

        return temp_excel

    except Exception as e:
        print(f"❌ Error generating Excel: {str(e)}")
        return None